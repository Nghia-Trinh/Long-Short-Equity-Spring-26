"""Alpha package API and command entrypoint for SUE + alpha artifacts.

This module serves two purposes:
1) Package surface (importable API used by other modules/scripts).
2) Executable script (`python Alpha/__init__.py`) that generates chart artifacts.
"""

# CLI parser used when this file is executed directly.
import argparse
# JSON loader used for reading config.json safely without pandas dtype coercion.
import json
# Regex used to normalize timestamped filenames when grouping chart families.
import re
# Dictionary helper for collecting files by group.
from collections import defaultdict
# Pathlib used for cross-platform filesystem operations.
from pathlib import Path

# Pandas for date parsing, CSV IO, and table/validation operations.
import pandas as pd
# NumPy for random ticker subset sampling.
import numpy as np

# Branch on package context to support both import mode and direct script mode.
if __package__:
    # Normal package mode (e.g., `import Alpha` from project root).
    from .alpha_matrix import (
        build_alpha_matrix,  # Core matrix constructor.
        plot_alpha_long_short_spread,  # Trend chart: long-short spread proxy.
        plot_alpha_matrix_snapshot,  # Table-like matrix chart.
        plot_alpha_dashboard,  # One-call alpha artifact generator.
        plot_alpha_heatmap,  # Heatmap chart for matrix values.
        plot_alpha_signal_strength,  # Line chart for average absolute alpha.
        save_alpha_matrix_csv,  # Matrix CSV export helper.
    )
    from .sue import (
        compute_sue,  # Core SUE computation.
        get_latest_sue_as_of,  # PIT SUE lookup.
        plot_sue_dashboard,  # One-call SUE chart generator.
        plot_sue_distribution,  # SUE histogram chart.
        plot_surprise_vs_returns,  # SUE vs price overlay chart.
    )
else:
    # Direct script fallback mode (e.g., running Alpha/__init__.py directly).
    # We inject project root into sys.path so `from Alpha...` imports resolve.
    import sys
    from pathlib import Path

    # Add repository root to the front of Python import search path.
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    # Re-import the same symbols through absolute package imports.
    from Alpha.alpha_matrix import (
        build_alpha_matrix,
        plot_alpha_long_short_spread,
        plot_alpha_matrix_snapshot,
        plot_alpha_dashboard,
        plot_alpha_heatmap,
        plot_alpha_signal_strength,
        save_alpha_matrix_csv,
    )
    from Alpha.sue import (
        compute_sue,
        get_latest_sue_as_of,
        plot_sue_dashboard,
        plot_sue_distribution,
        plot_surprise_vs_returns,
    )

# Explicit public API for `from Alpha import *` and tooling introspection.
__all__ = [
    "compute_sue",  # SUE batch computation.
    "get_latest_sue_as_of",  # PIT value lookup.
    "build_alpha_matrix",  # Alpha matrix constructor.
    "save_alpha_matrix_csv",  # Matrix saver.
    "plot_sue_distribution",  # SUE histogram chart.
    "plot_surprise_vs_returns",  # SUE vs price chart.
    "plot_sue_dashboard",  # SUE chart bundle.
    "plot_alpha_heatmap",  # Alpha heatmap.
    "plot_alpha_signal_strength",  # Alpha strength trend chart.
    "plot_alpha_long_short_spread",  # Long-short spread trend chart.
    "plot_alpha_matrix_snapshot",  # Alpha snapshot table chart.
    "plot_alpha_dashboard",  # Alpha chart bundle.
    "generate_artifact_pack",  # End-to-end artifact orchestrator.
    "export_artifact_pack_excel",  # Excel artifact exporter.
    "cleanup_chart_outputs",  # Chart dedup/cleanup utility.
    "cleanup_excel_outputs",  # Excel dedup/cleanup utility.
]


def _default_rebalance_dates(periods: int = 24) -> pd.DatetimeIndex:
    """Create default month-end rebalance dates from `Data/prices.csv`."""
    # Compute absolute path to prices file from project root.
    prices_path = Path(__file__).resolve().parents[1] / "Data" / "prices.csv"
    # Read only required column to reduce IO/memory cost.
    prices = pd.read_csv(prices_path, usecols=["trade_date"])
    # Parse dates and coerce malformed values to NaT.
    prices["trade_date"] = pd.to_datetime(prices["trade_date"], errors="coerce")
    # Convert daily dates to unique month-end timestamps, sorted ascending.
    month_ends = (
        prices["trade_date"]
        .dropna()
        .dt.to_period("M")
        .drop_duplicates()
        .dt.to_timestamp("M")
        .sort_values()
    )
    # Return empty index if no valid date could be formed.
    if month_ends.empty:
        return pd.DatetimeIndex([])
    # Keep only the most recent `periods` dates.
    return pd.DatetimeIndex(month_ends.tail(periods))


def _excel_dir() -> Path:
    """Return/create dedicated Excel output directory under outputs/."""
    excel_path = Path(__file__).resolve().parents[1] / "outputs" / "excel"
    excel_path.mkdir(parents=True, exist_ok=True)
    return excel_path


def _timestamped_excel_filename(base_name: str, timestamped: bool) -> str:
    """Build timestamped (or stable) .xlsx file names."""
    if not timestamped:
        return f"{base_name}.xlsx"
    stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_{stamp}.xlsx"


def _latest_artifact_files(chart_dir: Path) -> list[Path]:
    """Return newest artifact file per family (PNG and CSV tracked separately)."""
    grouped = defaultdict(list)
    for path in list(chart_dir.glob("*.png")) + list(chart_dir.glob("*.csv")) + list(chart_dir.glob("*.txt")):
        match = re.match(r"^(.*)_\d{8}_\d{6}$", path.stem)
        normalized_stem = match.group(1) if match else path.stem
        group_key = f"{normalized_stem}|{path.suffix.lower()}"
        grouped[group_key].append(path)

    latest = []
    for _, items in grouped.items():
        latest.append(sorted(items, key=lambda item: item.stat().st_mtime, reverse=True)[0])
    return sorted(latest, key=lambda item: item.name.lower())


def _latest_by_patterns(chart_dir: Path, patterns: list[str]) -> Path | None:
    """Return newest file matching first non-empty glob pattern sequence."""
    # Evaluate patterns in declared priority order and return first matching newest file.
    for pattern in patterns:
        matches = sorted(chart_dir.glob(pattern), key=lambda item: item.stat().st_mtime, reverse=True)
        if matches:
            return matches[0]
    # Return None when no pattern matches any file.
    return None


def _select_excel_chart_images(chart_dir: Path, representative_ticker: str) -> list[Path]:
    """Select a curated chart set for Excel Charts tab with one all-stocks heatmap."""
    # Build ordered chart selection candidates and keep at most one file per row.
    selected: list[Path] = []

    # Exactly one all-stocks heatmap: prefer explicit `alpha_heatmap_all*`, else latest base family excluding variant labels.
    all_heatmap = _latest_by_patterns(chart_dir, ["alpha_heatmap_all*.png"])
    if all_heatmap is None:
        all_heatmap_candidates = sorted(
            [
                path
                for path in chart_dir.glob("alpha_heatmap*.png")
                if "random20" not in path.stem and not path.stem.endswith("_80") and not path.stem.endswith("_all")
            ],
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        if all_heatmap_candidates:
            all_heatmap = all_heatmap_candidates[0]
    if all_heatmap is None:
        all_heatmap = _latest_by_patterns(chart_dir, ["alpha_heatmap.png"])
    if all_heatmap is not None:
        selected.append(all_heatmap)

    # Optional readability heatmap for random 20-ticker subset.
    random20_heatmap = _latest_by_patterns(chart_dir, ["alpha_heatmap_random20*.png"])
    if random20_heatmap is not None:
        selected.append(random20_heatmap)

    # Add core full-universe alpha diagnostics.
    for family_pattern in [
        "alpha_signal_strength*.png",
        "alpha_long_short_spread*.png",
        "alpha_matrix_snapshot*.png",
    ]:
        latest = _latest_by_patterns(chart_dir, [family_pattern])
        if latest is not None:
            selected.append(latest)

    # Add representative ticker SUE diagnostics.
    sue_distribution = _latest_by_patterns(chart_dir, [f"sue_distribution_{representative_ticker}*.png"])
    if sue_distribution is not None:
        selected.append(sue_distribution)
    sue_overlay = _latest_by_patterns(chart_dir, [f"sue_vs_price_{representative_ticker}*.png"])
    if sue_overlay is not None:
        selected.append(sue_overlay)

    # Remove accidental duplicates while preserving order.
    deduped: list[Path] = []
    seen = set()
    for chart_path in selected:
        if str(chart_path) in seen:
            continue
        seen.add(str(chart_path))
        deduped.append(chart_path)
    return deduped


def _build_chart_analysis_sections(
    alpha: pd.DataFrame,
    sue_df: pd.DataFrame,
    representative_ticker: str,
    random20_tickers: list[str] | None = None,
) -> list[dict]:
    """Build per-chart analysis sections (1-2 paragraphs per chart)."""
    date_start = pd.Timestamp(alpha.index.min()).strftime("%Y-%m-%d")
    date_end = pd.Timestamp(alpha.index.max()).strftime("%Y-%m-%d")
    ticker_count = int(alpha.shape[1])
    date_count = int(alpha.shape[0])

    alpha_mean_abs = alpha.abs().mean(axis=1)
    alpha_std = alpha.std(axis=1, ddof=0)
    long_short_proxy = alpha.apply(
        lambda row: (
            row.sort_values().iloc[-max(1, int(len(row) * 0.2)) :].mean()
            - row.sort_values().iloc[: max(1, int(len(row) * 0.2))].mean()
        )
        if len(row) > 0
        else 0.0,
        axis=1,
    )

    latest_row = alpha.iloc[-1]
    top_names = latest_row.sort_values(ascending=False).head(5).index.tolist()
    bottom_names = latest_row.sort_values(ascending=True).head(5).index.tolist()

    sue_ticker = sue_df[sue_df["ticker"] == representative_ticker].copy()
    latest_sue = float(sue_ticker["sue"].iloc[-1]) if not sue_ticker.empty else 0.0
    latest_surprise = float(sue_ticker["surprise"].iloc[-1]) if not sue_ticker.empty else 0.0
    sue_mean = float(sue_ticker["sue"].mean()) if not sue_ticker.empty else 0.0
    sue_std = float(sue_ticker["sue"].std(ddof=0)) if not sue_ticker.empty else 0.0
    sue_pos_rate = float((sue_ticker["sue"] > 0).mean()) if not sue_ticker.empty else 0.0
    sue_skew = float(sue_ticker["sue"].skew()) if not sue_ticker.empty else 0.0

    signal_slope = float(pd.Series(alpha_mean_abs.values).diff().mean())
    spread_mean = float(long_short_proxy.mean())
    spread_std = float(long_short_proxy.std(ddof=0))
    spread_pos_rate = float((long_short_proxy > 0).mean())

    return [
        {
            "chart": "SUE Distribution",
            "analysis": (
                f"This histogram summarizes standardized earnings surprises for {representative_ticker}. "
                f"Across available events, SUE has mean {sue_mean:.3f}, standard deviation {sue_std:.3f}, and skew {sue_skew:.3f}, "
                f"which indicates whether positive or negative outliers dominate. The positive-rate is {sue_pos_rate:.1%}, "
                f"so you can quickly assess directional balance in earnings signal realizations.\n\n"
                f"The most recent SUE is {latest_sue:.3f} with raw surprise {latest_surprise:.3f}. If this latest point lies in a tail "
                f"of the distribution relative to prior events, it generally reflects a higher-conviction earnings signal and is worth "
                f"cross-checking against subsequent price response and portfolio-level spread behavior."
            ),
        },
        {
            "chart": "SUE vs Price Overlay",
            "analysis": (
                f"This overlay compares event-timed SUE bars with the daily adjusted price path for {representative_ticker}. "
                f"Use it to visually test whether large positive/negative SUE observations are followed by directional price continuation "
                f"or reversal in the surrounding window.\n\n"
                f"Given latest SUE {latest_sue:.3f}, interpret the newest bar relative to price context: if extreme SUE values repeatedly align "
                f"with durable moves, that supports signal relevance; if responses are noisy or mean-reverting, it suggests weaker single-name "
                f"predictability and greater importance of cross-sectional portfolio construction."
            ),
        },
        {
            "chart": "Alpha Heatmap",
            "analysis": (
                f"The heatmap shows cross-sectional alpha values across {date_count} rebalance dates ({date_start} to {date_end}) and "
                f"{ticker_count} tickers. Dense color contrast indicates broad cross-sectional differentiation, which is desirable for "
                f"ranking and long-short selection.\n\n"
                f"Because rows are z-scored, each date is centered and scaled comparably. This means visual regime shifts mostly reflect changes in "
                f"relative ordering and dispersion, not level drift, making the chart useful for checking whether signal structure is stable over time."
            ),
        },
        {
            "chart": "Alpha Heatmap (Random 20)",
            "analysis": (
                f"This companion heatmap shows a random sample of {min(20, ticker_count)} tickers from the same alpha matrix window, "
                f"making single-name color patterns easier to inspect than the full-universe view. "
                f"Sample tickers include {', '.join((random20_tickers or [])[:10])}"
                f"{'' if not random20_tickers or len(random20_tickers) <= 10 else ', ...'}"
                f".\n\n"
                f"Use this view for visual spot-checks of row-wise cross-sectional differentiation and temporal consistency in ranking structure. "
                f"Because values remain z-scored by date, interpretation is directly comparable to the full heatmap while being more readable at the "
                f"individual ticker level."
            ),
        },
        {
            "chart": "Alpha Signal Strength",
            "analysis": (
                f"Signal-strength tracks mean absolute alpha per date. In this run, average mean-absolute-alpha is {float(alpha_mean_abs.mean()):.3f} "
                f"and average cross-sectional std is {float(alpha_std.mean()):.3f}, which indicates how strongly names are separated each rebalance.\n\n"
                f"The average first-difference trend of signal strength is {signal_slope:.4f} per step. Persistent decline can indicate crowding/noise, "
                f"while stable or rising values suggest the model continues to produce meaningful cross-sectional dispersion for portfolio formation."
            ),
        },
        {
            "chart": "Alpha Long/Short Spread",
            "analysis": (
                f"This chart measures top-20% minus bottom-20% alpha spread each rebalance as a proxy for long-short separation quality. "
                f"The spread averages {spread_mean:.3f} with std {spread_std:.3f}, and is positive on {spread_pos_rate:.1%} of dates.\n\n"
                f"Higher and stable spread suggests cleaner ranking separation between high-conviction longs and shorts. Compression or oscillation "
                f"in spread often flags weaker signal discrimination and should be monitored alongside transaction costs and turnover constraints."
            ),
        },
        {
            "chart": "Alpha Matrix Snapshot",
            "analysis": (
                f"The snapshot table provides an auditable slice of matrix values used by the optimizer and makes it easy to inspect exact entries. "
                f"On the latest date, top names include {', '.join(top_names)} and bottom names include {', '.join(bottom_names)} based on alpha ranking.\n\n"
                f"This table is especially useful for sanity checks (missing values, extreme outliers, sign consistency) and for validating that exported "
                f"data aligns with the plotted trends before portfolio optimization or reporting."
            ),
        },
    ]


def _compose_analysis_text(analysis_sections: list[dict]) -> str:
    """Compose a plain-text analysis report from per-chart sections."""
    blocks = []
    for section in analysis_sections:
        blocks.append(f"{section['chart']}\n{'-' * len(section['chart'])}\n{section['analysis']}")
    return "\n\n".join(blocks)


def _write_chart_analysis_file(charts_dir: Path, analysis_text: str, timestamped: bool) -> Path:
    """Write chart analysis text file into outputs/charts."""
    if timestamped:
        stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        file_path = charts_dir / f"chart_analysis_{stamp}.txt"
    else:
        file_path = charts_dir / "chart_analysis.txt"
    file_path.write_text(analysis_text, encoding="utf-8")
    return file_path


def _validate_artifact_outputs(
    alpha: pd.DataFrame,
    charts_dir: Path,
    excel_output: Path,
    analysis_output: Path,
    representative_ticker: str,
) -> dict:
    """Validate artifact completeness and basic consistency for accuracy checks."""
    if not excel_output.exists():
        raise ValueError(f"Expected Excel output not found: {excel_output}")
    if not analysis_output.exists():
        raise ValueError(f"Expected analysis output not found: {analysis_output}")

    required_chart_prefixes = [
        "alpha_heatmap",
        "alpha_heatmap_random20",
        "alpha_signal_strength",
        "alpha_long_short_spread",
        "alpha_matrix_snapshot",
        "alpha_matrix",
        f"sue_distribution_{representative_ticker}",
        f"sue_vs_price_{representative_ticker}",
        "chart_analysis",
    ]

    existing_names = [path.name for path in charts_dir.iterdir() if path.is_file()]
    missing_prefixes = []
    for prefix in required_chart_prefixes:
        if not any(name.startswith(prefix) for name in existing_names):
            missing_prefixes.append(prefix)
    if missing_prefixes:
        raise ValueError(f"Missing required artifact families: {', '.join(missing_prefixes)}")

    # Validate alpha_matrix CSV shape against in-memory matrix.
    alpha_csv_candidates = sorted(
        [path for path in charts_dir.glob("alpha_matrix*.csv")],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if not alpha_csv_candidates:
        raise ValueError("No alpha_matrix CSV artifact found")
    latest_alpha_csv = alpha_csv_candidates[0]
    alpha_csv_df = pd.read_csv(latest_alpha_csv, index_col=0)
    if alpha_csv_df.shape != alpha.shape:
        raise ValueError(
            f"Alpha CSV shape mismatch: csv={alpha_csv_df.shape}, expected={alpha.shape}"
        )

    return {
        "excel_exists": True,
        "analysis_exists": True,
        "required_families_present": True,
        "alpha_csv_matches_matrix_shape": True,
        "alpha_csv_path": str(latest_alpha_csv),
    }


def export_artifact_pack_excel(
    alpha: pd.DataFrame,
    sue_df: pd.DataFrame,
    run_summary: dict,
    analysis_sections: list[dict],
    timestamped: bool = True,
) -> Path:
    """Export run outputs to Excel with data sheets and embedded chart images."""
    output_path = _excel_dir() / _timestamped_excel_filename("alpha_artifact_pack", timestamped)
    chart_dir = Path(run_summary["charts_dir"])
    latest_artifacts = _latest_artifact_files(chart_dir)

    # Build summary key/value table.
    summary_df = pd.DataFrame([{"key": key, "value": str(value)} for key, value in run_summary.items()])
    analysis_df = pd.DataFrame(analysis_sections)

    # Build workbook guide so each tab has explicit purpose and key fields.
    workbook_guide_df = pd.DataFrame(
        [
            {
                "sheet": "RunSummary",
                "purpose": "Run configuration, ticker scope, and validation metadata.",
                "key_fields": "tickers_used_count, tickers_used_preview, data_scope_non_heatmap, heatmap_scope",
            },
            {
                "sheet": "TickersUsed",
                "purpose": "Complete list of tickers used in this run and whether they appear in Random20 heatmap.",
                "key_fields": "ticker, in_random20_heatmap",
            },
            {
                "sheet": "AlphaMatrix",
                "purpose": "Full alpha matrix (all selected tickers; all rebalance dates).",
                "key_fields": "Index=rebalance_date, columns=tickers, values=z-scored alpha",
            },
            {
                "sheet": "AlphaMatrixRandom20",
                "purpose": "Random 20-ticker readability subset from full alpha matrix.",
                "key_fields": "Index=rebalance_date, columns=random20 tickers",
            },
            {
                "sheet": "SUEDataSelected",
                "purpose": "SUE history restricted to tickers used in this run.",
                "key_fields": "ticker, report_date, surprise, sigma, sue",
            },
            {
                "sheet": "SUEDataAll",
                "purpose": "Complete SUE history across entire available universe.",
                "key_fields": "ticker, report_date, surprise, sigma, sue",
            },
            {
                "sheet": "KPI",
                "purpose": "Full-universe per-date diagnostics from AlphaMatrix.",
                "key_fields": "mean_alpha, mean_abs_alpha, std_alpha, long_short_spread_20pct",
            },
            {
                "sheet": "Artifacts",
                "purpose": "Latest output artifact manifest with timestamps.",
                "key_fields": "filename, type, path, last_modified",
            },
            {
                "sheet": "Analysis",
                "purpose": "Per-chart narrative interpretation and checks.",
                "key_fields": "chart, analysis",
            },
            {
                "sheet": "Charts",
                "purpose": "Embedded images for selected latest artifacts.",
                "key_fields": "Exactly one all-stocks heatmap + random20 heatmap + core diagnostics",
            },
        ]
    )

    # Keep SUE rows only for tickers used in current run.
    selected_tickers = run_summary.get("tickers_used", [])
    sue_subset = sue_df[sue_df["ticker"].isin(selected_tickers)].copy()
    if "report_date" in sue_subset.columns:
        sue_subset = sue_subset.sort_values(["ticker", "report_date"])

    # Also prepare full SUE table so all ticker history is visible in Excel.
    sue_all = sue_df.copy()
    if "report_date" in sue_all.columns:
        sue_all = sue_all.sort_values(["ticker", "report_date"])

    # Build KPI table aligned with project trend-monitoring objectives.
    alpha_clean = alpha.copy()
    kpi = pd.DataFrame(index=alpha_clean.index)
    kpi.index.name = "rebalance_date"
    kpi["mean_alpha"] = alpha_clean.mean(axis=1)
    kpi["mean_abs_alpha"] = alpha_clean.abs().mean(axis=1)
    kpi["std_alpha"] = alpha_clean.std(axis=1, ddof=0)

    # Per-date top/bottom 20% spread as a compact long-short signal-strength proxy.
    def _spread_20(row: pd.Series) -> float:
        clean = pd.to_numeric(row, errors="coerce").dropna()
        if clean.empty:
            return 0.0
        k = max(1, int(len(clean) * 0.2))
        ranked = clean.sort_values()
        return float(ranked.iloc[-k:].mean() - ranked.iloc[:k].mean())

    kpi["long_short_spread_20pct"] = alpha_clean.apply(_spread_20, axis=1)

    # Build artifact manifest table.
    artifact_rows = []
    for path in latest_artifacts:
        artifact_rows.append(
            {
                "filename": path.name,
                "type": path.suffix.lower(),
                "path": str(path),
                "last_modified": pd.Timestamp(path.stat().st_mtime, unit="s"),
            }
        )
    artifacts_df = pd.DataFrame(artifact_rows)

    # Build explicit ticker list table so universe membership is transparent.
    all_tickers_used = [str(ticker) for ticker in run_summary.get("tickers_used", [])]
    random20_tickers = set(str(ticker) for ticker in run_summary.get("random20_tickers", []))
    tickers_used_df = pd.DataFrame(
        {
            "ticker": all_tickers_used,
            "in_random20_heatmap": [ticker in random20_tickers for ticker in all_tickers_used],
        }
    )

    # Build fixed random-20 matrix table for Excel readability and review.
    random20_tickers = [str(ticker) for ticker in run_summary.get("random20_tickers", [])]
    random20_columns = [ticker for ticker in random20_tickers if ticker in alpha.columns]
    if not random20_columns:
        random20_columns = alpha.columns[: min(20, alpha.shape[1])].tolist()
    alpha_random20 = alpha.loc[:, random20_columns].copy()

    # Write tabular sheets first.
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        workbook_guide_df.to_excel(writer, sheet_name="WorkbookGuide", index=False)
        summary_df.to_excel(writer, sheet_name="RunSummary", index=False)
        tickers_used_df.to_excel(writer, sheet_name="TickersUsed", index=False)
        alpha.to_excel(writer, sheet_name="AlphaMatrix", index=True)
        alpha_random20.to_excel(writer, sheet_name="AlphaMatrixRandom20", index=True)
        sue_subset.to_excel(writer, sheet_name="SUEDataSelected", index=False)
        sue_all.to_excel(writer, sheet_name="SUEDataAll", index=False)
        kpi.to_excel(writer, sheet_name="KPI", index=True)
        artifacts_df.to_excel(writer, sheet_name="Artifacts", index=False)
        analysis_df.to_excel(writer, sheet_name="Analysis", index=False)

    # Then embed image artifacts in a dedicated Charts sheet.
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.formatting.rule import ColorScaleRule

        workbook = load_workbook(output_path)
        charts_sheet = workbook.create_sheet("Charts")
        charts_sheet["A1"] = "Latest chart images"
        charts_sheet["A2"] = "Charts tab includes exactly one all-stocks alpha heatmap per workbook."

        # Apply rainbow-style heatmap coloring to numeric alpha cells in AlphaMatrix.
        alpha_sheet = workbook["AlphaMatrix"]
        if alpha_sheet.max_row >= 2 and alpha_sheet.max_column >= 2:
            value_range = f"B2:{alpha_sheet.cell(row=alpha_sheet.max_row, column=alpha_sheet.max_column).coordinate}"
            alpha_sheet.conditional_formatting.add(
                value_range,
                ColorScaleRule(
                    start_type="min",
                    start_color="4B0082",  # Indigo (low)
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="00FF00",  # Green (mid)
                    end_type="max",
                    end_color="FF0000",  # Red (high)
                ),
            )

        # Apply the same rainbow-style heatmap coloring to random-20 alpha matrix sheet.
        alpha_random20_sheet = workbook["AlphaMatrixRandom20"]
        if alpha_random20_sheet.max_row >= 2 and alpha_random20_sheet.max_column >= 2:
            random20_value_range = (
                f"B2:{alpha_random20_sheet.cell(row=alpha_random20_sheet.max_row, column=alpha_random20_sheet.max_column).coordinate}"
            )
            alpha_random20_sheet.conditional_formatting.add(
                random20_value_range,
                ColorScaleRule(
                    start_type="min",
                    start_color="4B0082",  # Indigo (low)
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="00FF00",  # Green (mid)
                    end_type="max",
                    end_color="FF0000",  # Red (high)
                ),
            )

        # Populate analysis worksheet with wrapped text formatting.
        analysis_sheet = workbook["Analysis"]
        try:
            from openpyxl.styles import Alignment

            analysis_sheet.column_dimensions["A"].width = 28
            analysis_sheet.column_dimensions["B"].width = 165
            # Apply wrap formatting for each analysis row.
            for row_idx in range(2, analysis_sheet.max_row + 1):
                analysis_cell = analysis_sheet[f"B{row_idx}"]
                analysis_cell.alignment = Alignment(wrap_text=True, vertical="top")
                analysis_sheet.row_dimensions[row_idx].height = 170
        except Exception:
            pass

        anchor_row = 3
        representative_ticker = str(run_summary.get("ticker_for_sue_dashboard", "")).strip().upper()
        selected_chart_images = _select_excel_chart_images(chart_dir=chart_dir, representative_ticker=representative_ticker)
        for artifact in selected_chart_images:
            charts_sheet[f"A{anchor_row}"] = artifact.name
            image = XLImage(str(artifact))
            image.width = 960
            image.height = 540
            charts_sheet.add_image(image, f"A{anchor_row + 1}")
            anchor_row += 33

        workbook.save(output_path)
    except Exception:
        # If image embedding fails, workbook still contains all data sheets.
        pass

    return output_path


def _load_config() -> dict:
    """Load strategy config from project root `config.json`."""
    # Locate config in repository root.
    config_path = Path(__file__).resolve().parents[1] / "config.json"
    # If config does not exist, return defaults-only behavior.
    if not config_path.exists():
        return {}
    # Read UTF-8 JSON into plain Python dict.
    with config_path.open("r", encoding="utf-8") as file:
        return json.load(file)


def _default_universe_from_prices(
    exclude_top_n_largecap: int = 500,
    min_dollar_volume: float = 1_000_000,
    max_tickers: int | None = None,
) -> list[str]:
    """Approximate investable universe from `prices.csv` using dollar-volume rules.

    Logic mirrors project design:
    1) Median trailing dollar volume (252 rows) per ticker.
    2) Remove top-N most-liquid names (large-cap proxy).
    3) Keep names above minimum dollar-volume threshold.
    """
    # Build path to source prices file.
    prices_path = Path(__file__).resolve().parents[1] / "Data" / "prices.csv"
    # Read only fields needed for dollar-volume screening.
    prices = pd.read_csv(prices_path, usecols=["ticker", "trade_date", "adj_close", "volume"])

    # Normalize ticker symbols to uppercase canonical form.
    prices["ticker"] = prices["ticker"].astype(str).str.strip().str.upper()
    # Parse trade date field for chronological operations.
    prices["trade_date"] = pd.to_datetime(prices["trade_date"], errors="coerce")
    # Coerce prices to float; malformed values become NaN.
    prices["adj_close"] = pd.to_numeric(prices["adj_close"], errors="coerce")
    # Coerce volume to numeric for multiplication.
    prices["volume"] = pd.to_numeric(prices["volume"], errors="coerce")
    # Drop rows missing required data for dollar-volume calculation.
    prices = prices.dropna(subset=["ticker", "trade_date", "adj_close", "volume"])
    # If nothing usable, return empty list.
    if prices.empty:
        return []

    # Ensure rows are ordered so trailing selection is meaningful per ticker.
    prices = prices.sort_values(["ticker", "trade_date"])
    # Compute daily dollar volume proxy.
    prices["dollar_volume"] = prices["adj_close"] * prices["volume"]

    # Keep trailing 252 rows (about 1 year of trading days) per ticker.
    trailing = prices.groupby("ticker", group_keys=False).tail(252)
    # Aggregate to median dollar volume, then rank descending by liquidity.
    median_dv = trailing.groupby("ticker")["dollar_volume"].median().sort_values(ascending=False)
    # If no groups survived, exit early.
    if median_dv.empty:
        return []

    # Exclude top liquid names to proxy out large caps.
    candidates = median_dv.iloc[max(0, int(exclude_top_n_largecap)) :]
    # Enforce minimum liquidity floor.
    candidates = candidates[candidates >= float(min_dollar_volume)]
    # Return empty if screen is too strict.
    if candidates.empty:
        return []

    # Return full ticker list (or capped list when requested).
    ticker_list = candidates.index.astype(str).tolist()
    if max_tickers is None:
        return ticker_list
    return ticker_list[:max_tickers]


def _validate_alpha_output(alpha: pd.DataFrame) -> dict:
    """Run numerical sanity checks to ensure generated alpha is coherent."""
    # Matrix must contain at least one row/column.
    if alpha.empty:
        raise ValueError("Generated alpha matrix is empty")
    # Date index must not contain duplicates.
    if alpha.index.has_duplicates:
        raise ValueError("Alpha matrix index contains duplicate dates")
    # Date index should be strictly sorted ascending.
    if not alpha.index.is_monotonic_increasing:
        raise ValueError("Alpha matrix index is not sorted increasing by date")

    # Convert to dense float array for fast finite checks.
    values = alpha.to_numpy(dtype=float)
    # Reject NaN values.
    if not pd.notna(values).all():
        raise ValueError("Alpha matrix contains NaN values")
    # Reject +inf/-inf values.
    if not ((values != float("inf")) & (values != float("-inf"))).all():
        raise ValueError("Alpha matrix contains infinite values")

    # Row-wise z-score output should be centered around zero.
    row_mean_max_abs = float(alpha.mean(axis=1).abs().max())
    # Use tight tolerance for mean-centering.
    if row_mean_max_abs > 1e-6:
        raise ValueError(f"Alpha row means are not centered near zero (max abs mean={row_mean_max_abs:.6g})")

    # Compute row-wise std (population std) expected from z-scoring.
    row_std = alpha.std(axis=1, ddof=0)
    # Valid rows are either std≈1 (normal) or std≈0 (constant row fallback).
    invalid_std = ~((row_std - 1.0).abs() < 1e-6) & ~(row_std.abs() < 1e-12)
    # Fail fast if any invalid std row exists.
    if bool(invalid_std.any()):
        raise ValueError("Alpha row standard deviation check failed for one or more dates")

    # Return a compact validation summary for logs/CLI output.
    return {
        "rows": int(alpha.shape[0]),  # Number of rebalance dates.
        "cols": int(alpha.shape[1]),  # Number of tickers.
        "max_abs_row_mean": row_mean_max_abs,  # Centering quality metric.
        "constant_rows": int((row_std.abs() < 1e-12).sum()),  # Rows with zero dispersion.
    }


def generate_artifact_pack(
    tickers: list[str] | None = None,
    lookback_quarters: int = 8,
    rebalance_periods: int = 24,
    max_tickers: int | None = None,
    use_ic_weighting: bool = True,
    show: bool = False,
    timestamped: bool = True,
    cleanup_replaced_charts: bool = True,
    keep_latest_charts: int = 1,
    keep_latest_excel: int = 1,
    min_default_universe_size: int = 25,
    remove_irrelevant_ticker_charts: bool = True,
) -> dict:
    """Generate SUE + alpha artifact pack into `outputs/charts`.

    Steps:
    - Compute SUE table and choose a representative ticker for SUE charts.
    - Build alpha artifacts across recent rebalance dates.
    - Validate generated alpha matrix numerics.
    - Optionally cleanup replaced chart/csv versions.
    """
    # Load config-driven defaults (liquidity filters, etc.).
    config = _load_config()
    # Read top-N exclusion parameter with fallback.
    default_exclude_top_n = int(config.get("universe_exclude_top_n_largecap", 500))
    # Read minimum dollar volume threshold with fallback.
    default_min_dv = float(config.get("universe_min_dollar_volume", 1_000_000))

    # Compute SUE panel once; reused for ticker availability filtering.
    sue_df = compute_sue(lookback_quarters=lookback_quarters)
    # Extract sorted unique tickers present in SUE data.
    available_tickers = sorted(sue_df["ticker"].dropna().unique().tolist())
    # Abort if no tickers are available.
    if not available_tickers:
        raise ValueError("No tickers found in earnings data; cannot generate artifacts")

    # If caller did not pass tickers, default to ALL available SUE tickers unless capped.
    if tickers is None:
        selected_tickers = available_tickers if max_tickers is None else available_tickers[:max_tickers]
    else:
        # Normalize user-provided tickers to uppercase canonical form.
        selected_tickers = [str(ticker).strip().upper() for ticker in tickers if str(ticker).strip()]
        # Fallback if caller passed only blanks/invalid entries.
        if not selected_tickers:
            selected_tickers = available_tickers if max_tickers is None else available_tickers[:max_tickers]

    # Precompute set for O(1) membership checks.
    available_set = set(available_tickers)
    # Remove any ticker missing from computed SUE universe.
    selected_tickers = [ticker for ticker in selected_tickers if ticker in available_set]
    # Final fallback if filtering removed everything.
    if not selected_tickers:
        selected_tickers = available_tickers if max_tickers is None else available_tickers[:max_tickers]

    # Keep compatibility for callers that still pass this parameter.
    if tickers is None and len(selected_tickers) < min_default_universe_size:
        selected_tickers = available_tickers if max_tickers is None else available_tickers[:max_tickers]

    # Use first selected ticker as representative SUE visualization symbol.
    representative_ticker = selected_tickers[0]
    # Generate SUE charts (distribution + SUE vs price).
    plot_sue_dashboard(
        ticker=representative_ticker,
        show=show,
        timestamped=timestamped,
    )

    # Build recent month-end rebalance timeline.
    rebalance_dates = _default_rebalance_dates(periods=rebalance_periods)
    # Abort if we cannot form any rebalance dates.
    if rebalance_dates.empty:
        raise ValueError("No rebalance dates available from prices.csv; cannot generate alpha artifacts")

    # Generate complete alpha artifact bundle (CSV + charts).
    alpha = plot_alpha_dashboard(
        rebalance_dates=rebalance_dates,
        tickers=selected_tickers,
        lookback_quarters=lookback_quarters,
        use_ic_weighting=use_ic_weighting,
        show=show,
        timestamped=timestamped,
    )

    # Build one extra readability heatmap from a random 20-ticker subset.
    random20_count = min(20, alpha.shape[1])
    random20_tickers = np.random.default_rng().choice(alpha.columns.to_numpy(), size=random20_count, replace=False).tolist()
    random20_alpha = alpha.loc[:, random20_tickers]
    chart_dir = Path(__file__).resolve().parents[1] / "outputs" / "charts"
    if timestamped:
        stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        random20_filename = f"alpha_heatmap_random20_{stamp}.png"
    else:
        random20_filename = "alpha_heatmap_random20.png"
    plot_alpha_heatmap(
        alpha_matrix=random20_alpha,
        max_tickers=None,
        show=show,
        save=True,
        save_path=chart_dir / random20_filename,
        timestamped=timestamped,
    )

    # Validate final alpha matrix for numerical integrity.
    validation = _validate_alpha_output(alpha)

    # Initialize cleanup summary to None when cleanup is disabled.
    cleanup_summary = None

    # Build base summary payload before Excel export so the workbook can include it.
    run_summary = {
        "charts_dir": str(Path(__file__).resolve().parents[1] / "outputs" / "charts"),
        "ticker_for_sue_dashboard": representative_ticker,
        "alpha_shape": tuple(alpha.shape),
        "tickers_used": selected_tickers,
        "tickers_used_count": int(len(selected_tickers)),
        "tickers_used_preview": selected_tickers[:20],
        "random20_tickers": random20_tickers,
        "data_scope_non_heatmap": "All selected tickers and all rebalance dates",
        "heatmap_scope": "Alpha Heatmap=all selected tickers; Alpha Heatmap (Random20)=20-ticker readability subset",
        "rebalance_dates": len(rebalance_dates),
        "validation": validation,
        "cleanup": cleanup_summary,
    }

    # Export an Excel workbook in outputs/excel containing data + embedded charts.
    analysis_sections = _build_chart_analysis_sections(
        alpha=alpha,
        sue_df=sue_df,
        representative_ticker=representative_ticker,
        random20_tickers=random20_tickers,
    )
    analysis_text = _compose_analysis_text(analysis_sections)
    analysis_file = _write_chart_analysis_file(
        charts_dir=Path(run_summary["charts_dir"]),
        analysis_text=analysis_text,
        timestamped=timestamped,
    )

    excel_output = export_artifact_pack_excel(
        alpha=alpha,
        sue_df=sue_df,
        run_summary=run_summary,
        analysis_sections=analysis_sections,
        timestamped=timestamped,
    )

    # Run final chart cleanup after all artifacts (including analysis txt) are written.
    if cleanup_replaced_charts:
        cleanup_summary = cleanup_chart_outputs(keep_latest=keep_latest_charts, dry_run=False)

        # Remove ticker-specific SUE charts not tied to current representative ticker.
        if remove_irrelevant_ticker_charts:
            chart_dir = Path(__file__).resolve().parents[1] / "outputs" / "charts"
            stale_ticker_files = []
            for stale_path in list(chart_dir.glob("sue_distribution_*.png")) + list(chart_dir.glob("sue_vs_price_*.png")):
                if representative_ticker not in stale_path.name:
                    stale_ticker_files.append(stale_path)
            for stale_path in stale_ticker_files:
                stale_path.unlink(missing_ok=True)
            cleanup_summary["removed_irrelevant_ticker_charts"] = [str(path) for path in stale_ticker_files]

    # Store final chart cleanup result in run summary.
    run_summary["cleanup"] = cleanup_summary

    # Keep only latest N Excel workbooks.
    excel_cleanup = cleanup_excel_outputs(keep_latest=keep_latest_excel, dry_run=False)

    # Validate final artifacts for expected formatting/completeness and consistency.
    artifact_validation = _validate_artifact_outputs(
        alpha=alpha,
        charts_dir=Path(run_summary["charts_dir"]),
        excel_output=excel_output,
        analysis_output=analysis_file,
        representative_ticker=representative_ticker,
    )

    # Return structured run summary for logs and automation hooks.
    run_summary["excel_output"] = str(excel_output)
    run_summary["analysis_output"] = str(analysis_file)
    run_summary["excel_cleanup"] = excel_cleanup
    run_summary["artifact_validation"] = artifact_validation
    return run_summary


def cleanup_chart_outputs(
    keep_latest: int = 3,
    chart_dir: str | Path | None = None,
    dry_run: bool = False,
) -> dict:
    """Remove older chart/csv artifacts while keeping newest N per family.

    Grouping behavior:
    - Timestamped files like `alpha_heatmap_YYYYMMDD_HHMMSS.png` are grouped as `alpha_heatmap|.png`.
    - Non-timestamped files are grouped by their full stem + extension.
    """
    # Guard against invalid negative retention request.
    if keep_latest < 0:
        raise ValueError("keep_latest must be >= 0")

    # Resolve target directory (default: outputs/charts).
    resolved_chart_dir = Path(chart_dir) if chart_dir is not None else Path(__file__).resolve().parents[1] / "outputs" / "charts"
    # If directory does not exist yet, return empty summary.
    if not resolved_chart_dir.exists():
        return {
            "chart_dir": str(resolved_chart_dir),  # Directory inspected.
            "removed": 0,  # Removed file count.
            "kept": 0,  # Kept file count.
            "removed_files": [],  # Removed file list.
        }

    # Initialize file groups keyed by normalized family name + extension.
    grouped_files = defaultdict(list)
    # Include chart images, matrix CSV artifacts, and text analysis files.
    chart_files = list(resolved_chart_dir.glob("*.png")) + list(resolved_chart_dir.glob("*.csv")) + list(resolved_chart_dir.glob("*.txt"))
    # Iterate all candidate files and place into their groups.
    for chart_path in chart_files:
        # Match optional timestamp suffix for normalization.
        match = re.match(r"^(.*)_\d{8}_\d{6}$", chart_path.stem)
        # If timestamp pattern exists, strip it; otherwise use raw stem.
        normalized_stem = match.group(1) if match else chart_path.stem
        # Keep extension in group key so PNG and CSV lifecycles remain independent.
        group_key = f"{normalized_stem}|{chart_path.suffix.lower()}"
        # Append file to the computed group.
        grouped_files[group_key].append(chart_path)

    # Accumulator for removed file paths.
    removed_files = []
    # Counter for kept files across all groups.
    kept_count = 0
    # Process each group independently.
    for _, file_group in grouped_files.items():
        # Sort newest first by modification time.
        ordered = sorted(file_group, key=lambda item: item.stat().st_mtime, reverse=True)
        # Count files we are retaining from the front of sorted list.
        kept_count += len(ordered[:keep_latest])
        # Iterate stale tail files beyond retention limit.
        for stale_file in ordered[keep_latest:]:
            # Record file path in summary.
            removed_files.append(str(stale_file))
            # Delete file unless dry-run mode is enabled.
            if not dry_run:
                # `missing_ok` avoids failure in case file was concurrently removed.
                stale_file.unlink(missing_ok=True)

    # Return cleanup summary payload.
    return {
        "chart_dir": str(resolved_chart_dir),  # Directory cleaned.
        "removed": len(removed_files),  # Number of files removed.
        "kept": kept_count,  # Number of files retained.
        "removed_files": removed_files,  # Exact removed file paths.
    }


def cleanup_excel_outputs(
    keep_latest: int = 1,
    excel_dir: str | Path | None = None,
    dry_run: bool = False,
) -> dict:
    """Remove older Excel artifact packs, keeping newest N files."""
    if keep_latest < 0:
        raise ValueError("keep_latest must be >= 0")

    resolved_excel_dir = Path(excel_dir) if excel_dir is not None else _excel_dir()
    if not resolved_excel_dir.exists():
        return {
            "excel_dir": str(resolved_excel_dir),
            "removed": 0,
            "kept": 0,
            "removed_files": [],
        }

    workbook_files = sorted(
        resolved_excel_dir.glob("alpha_artifact_pack*.xlsx"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    keep_files = workbook_files[:keep_latest]
    stale_files = workbook_files[keep_latest:]

    removed_files = []
    skipped_files = []
    for stale_file in stale_files:
        stale_path_str = str(stale_file)
        if not dry_run:
            try:
                stale_file.unlink(missing_ok=True)
                removed_files.append(stale_path_str)
            except PermissionError:
                # File is open/locked (common on Windows when workbook is open in Excel).
                skipped_files.append(stale_path_str)
        else:
            removed_files.append(stale_path_str)

    return {
        "excel_dir": str(resolved_excel_dir),
        "removed": len(removed_files),
        "kept": len(keep_files),
        "removed_files": removed_files,
        "skipped_locked_files": skipped_files,
    }


# Execute CLI workflow only when file is run directly, not when imported.
if __name__ == "__main__":
    # Create parser for script-level options.
    parser = argparse.ArgumentParser(description="Generate Alpha SUE and matrix artifact pack")
    # Optional user ticker list as comma-separated values.
    parser.add_argument(
        "--tickers",
        type=str,
        default="",
        help="Comma-separated ticker list (example: A,ABT,ACN). Defaults to first 25 available.",
    )
    # SUE rolling lookback parameter.
    parser.add_argument("--lookback-quarters", type=int, default=8, help="SUE rolling lookback window in quarters")
    # Number of month-end rebalance points.
    parser.add_argument("--rebalance-periods", type=int, default=24, help="Number of recent month-end rebalance dates")
    # Max tickers when default universe generation is used.
    parser.add_argument("--max-tickers", type=int, default=0, help="Maximum number of tickers when default universe is used (0 means all)")
    parser.add_argument("--all-tickers", action="store_true", help="Use all available filtered tickers")
    # Display charts interactively if desired.
    parser.add_argument("--show", dest="show", action="store_true", help="Display plots interactively")
    # Explicitly disable interactive display (alias-friendly counterpart to --show).
    parser.add_argument("--no-show", dest="show", action="store_false", help="Do not display plots interactively")
    # Default behavior is non-interactive output generation.
    parser.set_defaults(show=False)
    # Disable IC weighting switch.
    parser.add_argument("--no-ic-weighting", action="store_true", help="Disable IC weighting in alpha matrix")
    # Disable timestamped filenames switch.
    parser.add_argument("--no-timestamp", action="store_true", help="Disable timestamped output filenames")
    # Keep-latest retention count for cleanup.
    parser.add_argument(
        "--keep-latest-charts",
        type=int,
        default=1,
        help="Keep latest N chart files per chart family after generation",
    )
    # Keep-latest retention count for Excel artifact workbooks.
    parser.add_argument(
        "--keep-latest-excel",
        type=int,
        default=1,
        help="Keep latest N Excel artifact workbooks after generation",
    )
    # Minimum universe size before default filter broadens to available SUE names.
    parser.add_argument(
        "--min-default-universe-size",
        type=int,
        default=25,
        help="If default filtered universe is smaller than this, broaden to available SUE universe",
    )
    # Allow disabling automatic cleanup.
    parser.add_argument("--no-cleanup", action="store_true", help="Do not delete older duplicate chart files")
    # Preserve legacy SUE ticker chart images instead of removing non-current tickers.
    parser.add_argument("--keep-irrelevant-ticker-charts", action="store_true", help="Keep SUE charts for tickers not used in current run")

    # Parse CLI args.
    args = parser.parse_args()
    # Parse ticker csv string into normalized list.
    ticker_list = [ticker.strip().upper() for ticker in args.tickers.split(",") if ticker.strip()]
    # Convert CLI settings to optional cap semantics.
    effective_max_tickers = None if args.all_tickers or args.max_tickers == 0 else int(args.max_tickers)

    # Run artifact generation with parsed CLI options.
    summary = generate_artifact_pack(
        tickers=ticker_list or None,
        lookback_quarters=args.lookback_quarters,
        rebalance_periods=args.rebalance_periods,
        max_tickers=effective_max_tickers,
        use_ic_weighting=not args.no_ic_weighting,
        show=args.show,
        timestamped=not args.no_timestamp,
        cleanup_replaced_charts=not args.no_cleanup,
        keep_latest_charts=args.keep_latest_charts,
        keep_latest_excel=args.keep_latest_excel,
        min_default_universe_size=args.min_default_universe_size,
        remove_irrelevant_ticker_charts=not args.keep_irrelevant_ticker_charts,
    )
    # Print completion message.
    print("Alpha artifact pack generated.")
    # Print structured run summary.
    print(summary)
